"""Evaluation report serialization and console output."""

from dataclasses import dataclass, field
from typing import Dict, List, Any
from pathlib import Path
import csv
import json

from rich.console import Console
from rich.table import Table


from ticket_router_base.data import BaseDataset

from .evaluator import TaskEvaluationResult

console = Console()


@dataclass
class EvaluationReport:
    """Complete evaluation report for a model across all dataset tasks."""

    model_name: str
    dataset: BaseDataset

    file_path: Path | None
    task_results: List[TaskEvaluationResult]

    error_summary: Dict[str, int]  # error flag name -> count
    total_samples: int

    # Multi-run aggregation fields
    task_stds: List[TaskEvaluationResult] = field(default_factory=list)
    run_results: List[List[TaskEvaluationResult]] = field(default_factory=list)
    n_runs: int = 1

    # Config discovery fields
    cfg_id: str | None = None
    cfg_info: Dict[str, Any] | None = None


def print_overall_report(reports: List[EvaluationReport]) -> None:
    if not reports:
        console.print("[red]No reports provided[/red]")
        return

    dataset_names = set(r.dataset.name for r in reports)
    if len(dataset_names) != 1:
        raise ValueError(f"Reports must share same dataset: {dataset_names}")

    all_tasks = sorted({t.task_name for r in reports for t in r.task_results})

    fairness_keys = sorted(
        {key for r in reports for t in r.task_results for key in t.fairness.keys()}
    )

    def get_task(
        report: EvaluationReport, task_name: str
    ) -> TaskEvaluationResult | None:
        return next(
            (t for t in report.task_results if t.task_name == task_name),
            None,
        )

    def get_std_task(
        report: EvaluationReport, task_name: str
    ) -> TaskEvaluationResult | None:
        if not report.task_stds:
            return None
        return next(
            (t for t in report.task_stds if t.task_name == task_name),
            None,
        )

    def format_metric(val: float | None, std: float | None) -> str:
        if val is None:
            return "-"
        if std is not None and std > 0:
            return f"{val:.4f} ± {std:.4f}"
        return f"{val:.4f}"

    for task_name in all_tasks:
        table = Table(
            title=f"Task: {task_name}",
            show_lines=False,
        )

        table.add_column("Metric")

        for r in reports:
            if r.cfg_id is not None and r.cfg_info:
                # Show first 2 config entries as summary
                cfg_summary = ", ".join(
                    f"{k}={v}" for k, v in list(r.cfg_info.items())[:2]
                )
                # Escape '[' for rich markup; wrap in parens instead of brackets
                col_name = f"{r.model_name}\n{r.cfg_id} (n={r.n_runs})\n({cfg_summary})"
            elif r.n_runs > 1:
                col_name = f"{r.model_name} (n={r.n_runs})"
            else:
                col_name = r.model_name
            table.add_column(col_name, justify="right")

        base_metrics = [
            ("Accuracy", "accuracy"),
            ("Macro F1", "macro_f1"),
            ("MAE", "mae"),
            ("QWK", "qwk"),
        ]

        for metric_name, metric_key in base_metrics:
            row = [metric_name]

            for r in reports:
                task = get_task(r, task_name)
                if task is None:
                    row.append("-")
                    continue

                val = getattr(task.performance, metric_key, None)
                std_task = get_std_task(r, task_name)
                std = getattr(std_task.performance, metric_key, None) if std_task is not None else None
                row.append(format_metric(val, std))

            table.add_row(*row)

        fairness_metrics = [
            ("Acc Gap", "accuracy_gap"),
            ("Acc Ratio", "accuracy_ratio"),
            ("F1 Gap", "macro_f1_gap"),
            ("F1 Ratio", "macro_f1_ratio"),
            ("DI", "avg_disparate_impact"),
        ]

        for key in fairness_keys:
            for metric_name, field_name in fairness_metrics:
                row = [f"{key} {metric_name}"]

                for r in reports:
                    task = get_task(r, task_name)
                    if task is None:
                        row.append("-")
                        continue

                    fm = task.fairness.get(key)
                    if fm is None:
                        row.append("-")
                        continue

                    val = getattr(fm, field_name, None)
                    std_task = get_std_task(r, task_name)
                    std_fm = std_task.fairness.get(key) if std_task is not None else None
                    std = getattr(std_fm, field_name, None) if std_fm is not None else None
                    row.append(format_metric(val, std))

                table.add_row(*row)

        console.print(table)


def _get_task(report: EvaluationReport, task_name: str) -> TaskEvaluationResult | None:
    return next(
        (t for t in report.task_results if t.task_name == task_name),
        None,
    )


def _get_std_task(report: EvaluationReport, task_name: str) -> TaskEvaluationResult | None:
    if not report.task_stds:
        return None
    return next(
        (t for t in report.task_stds if t.task_name == task_name),
        None,
    )


def _build_tidy_rows(reports: List[EvaluationReport]) -> List[Dict[str, Any]]:
    """Build tidy-format rows from evaluation reports (shared by CSV and Excel)."""
    if not reports:
        return []

    all_tasks = sorted({t.task_name for r in reports for t in r.task_results})
    fairness_keys = sorted(
        {key for r in reports for t in r.task_results for key in t.fairness.keys()}
    )

    rows: List[Dict[str, Any]] = []

    for task_name in all_tasks:
        perf_fields = [
            ("accuracy", "accuracy"),
            ("macro_precision", "macro_precision"),
            ("macro_recall", "macro_recall"),
            ("macro_f1", "macro_f1"),
            ("mae", "mae"),
            ("qwk", "qwk"),
        ]

        for metric_name, attr_name in perf_fields:
            for r in reports:
                task = _get_task(r, task_name)
                if task is None:
                    continue
                val = getattr(task.performance, attr_name, None)
                if val is None:
                    continue

                std_task = _get_std_task(r, task_name)
                std = getattr(std_task.performance, attr_name, None) if std_task is not None else None

                rows.append({
                    "task_name": task_name,
                    "model_name": r.model_name,
                    "cfg": json.dumps(r.cfg_info, ensure_ascii=False, sort_keys=True) if r.cfg_info else "",
                    "n_runs": r.n_runs,
                    "metric_category": "performance",
                    "metric_name": metric_name,
                    "sensitive_attr": "",
                    "value": val,
                    "std": std if std is not None else "",
                })

        fairness_fields = [
            ("accuracy_gap", "accuracy_gap"),
            ("accuracy_ratio", "accuracy_ratio"),
            ("macro_f1_gap", "macro_f1_gap"),
            ("macro_f1_ratio", "macro_f1_ratio"),
            ("avg_disparate_impact", "avg_disparate_impact"),
            ("avg_equal_opportunity_difference", "avg_equal_opportunity_difference"),
            ("avg_average_odds_difference", "avg_average_odds_difference"),
        ]

        for sensitive_attr in fairness_keys:
            for metric_name, attr_name in fairness_fields:
                for r in reports:
                    task = _get_task(r, task_name)
                    if task is None:
                        continue
                    fm = task.fairness.get(sensitive_attr)
                    if fm is None:
                        continue
                    val = getattr(fm, attr_name, None)
                    if val is None:
                        continue

                    std_task = _get_std_task(r, task_name)
                    std_fm = std_task.fairness.get(sensitive_attr) if std_task is not None else None
                    std = getattr(std_fm, attr_name, None) if std_fm is not None else None

                    rows.append({
                        "task_name": task_name,
                        "model_name": r.model_name,
                        "cfg": json.dumps(r.cfg_info, ensure_ascii=False, sort_keys=True) if r.cfg_info else "",
                        "n_runs": r.n_runs,
                        "metric_category": "fairness",
                        "metric_name": metric_name,
                        "sensitive_attr": sensitive_attr,
                        "value": val,
                        "std": std if std is not None else "",
                    })

    return rows


def save_reports_to_csv(reports: List[EvaluationReport], output_path: Path) -> None:
    """Save evaluation reports to a tidy CSV file."""
    rows = _build_tidy_rows(reports)
    if not rows:
        return

    fieldnames = [
        "task_name",
        "model_name",
        "cfg",
        "n_runs",
        "metric_category",
        "metric_name",
        "sensitive_attr",
        "value",
        "std",
    ]

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    console.print(f"[green]CSV saved to {output_path} ({len(rows)} rows)[/green]")


def _format_metric(val: float | None, std: float | None) -> str:
    if val is None:
        return "-"
    if std is not None and std > 0:
        return f"{val:.4f} ± {std:.4f}"
    return f"{val:.4f}"


def _write_overview_sheet(
    ws,
    reports: List[EvaluationReport],
    task_name: str,
    cfg_keys: List[str],
    fairness_keys: List[str],
) -> None:
    """Write a single task overview sheet using openpyxl with multi-index columns."""
    from openpyxl.styles import Font, Alignment

    n_reports = len(reports)
    n_header_rows = 1 + len(cfg_keys)  # model_name + cfg keys
    header_offset = n_header_rows + 1  # +1 blank row

    # Column header: row 1 = model_name, row 2..n = cfg values
    for col_idx, r in enumerate(reports, start=2):  # col 1 reserved for metric names
        ws.cell(row=1, column=col_idx, value=r.model_name)
        for k_idx, key in enumerate(cfg_keys, start=2):
            val = r.cfg_info.get(key, "") if r.cfg_info else ""
            ws.cell(row=k_idx, column=col_idx, value=str(val))

    # Merge contiguous model_name cells
    if n_reports > 1:
        start_col = 2
        current_model = reports[0].model_name
        for col_idx in range(3, n_reports + 2):
            if reports[col_idx - 2].model_name != current_model:
                if col_idx - 1 > start_col:
                    ws.merge_cells(
                        start_row=1, start_column=start_col,
                        end_row=1, end_column=col_idx - 1
                    )
                start_col = col_idx
                current_model = reports[col_idx - 2].model_name
        if n_reports + 1 > start_col:
            ws.merge_cells(
                start_row=1, start_column=start_col,
                end_row=1, end_column=n_reports + 1
            )

    # Style header rows
    for row_idx in range(1, n_header_rows + 1):
        for col_idx in range(1, n_reports + 2):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Metric rows
    metric_rows: List[tuple[str, List[str]]] = []

    perf_fields = [
        ("accuracy", "accuracy"),
        ("macro_precision", "macro_precision"),
        ("macro_recall", "macro_recall"),
        ("macro_f1", "macro_f1"),
        ("mae", "mae"),
        ("qwk", "qwk"),
    ]

    for metric_label, attr_name in perf_fields:
        values: List[str] = []
        for r in reports:
            task = _get_task(r, task_name)
            if task is None:
                values.append("-")
                continue
            val = getattr(task.performance, attr_name, None)
            std_task = _get_std_task(r, task_name)
            std = getattr(std_task.performance, attr_name, None) if std_task is not None else None
            values.append(_format_metric(val, std))
        metric_rows.append((metric_label, values))

    fairness_fields = [
        ("accuracy_gap", "accuracy_gap"),
        ("accuracy_ratio", "accuracy_ratio"),
        ("macro_f1_gap", "macro_f1_gap"),
        ("macro_f1_ratio", "macro_f1_ratio"),
        ("avg_disparate_impact", "avg_disparate_impact"),
        ("avg_equal_opportunity_difference", "avg_equal_opportunity_difference"),
        ("avg_average_odds_difference", "avg_average_odds_difference"),
    ]

    for sensitive_attr in fairness_keys:
        for metric_label, attr_name in fairness_fields:
            values: List[str] = []
            for r in reports:
                task = _get_task(r, task_name)
                if task is None:
                    values.append("-")
                    continue
                fm = task.fairness.get(sensitive_attr)
                if fm is None:
                    values.append("-")
                    continue
                val = getattr(fm, attr_name, None)
                std_task = _get_std_task(r, task_name)
                std_fm = std_task.fairness.get(sensitive_attr) if std_task is not None else None
                std = getattr(std_fm, attr_name, None) if std_fm is not None else None
                values.append(_format_metric(val, std))
            metric_rows.append((f"{sensitive_attr} {metric_label}", values))

    for row_offset, (metric_label, values) in enumerate(metric_rows, start=1):
        row_idx = header_offset + row_offset
        ws.cell(row=row_idx, column=1, value=metric_label)
        for col_idx, val in enumerate(values, start=2):
            ws.cell(row=row_idx, column=col_idx, value=val)

    # Auto-adjust column width for metric names
    ws.column_dimensions["A"].width = max(len(label) for label, _ in metric_rows) + 4


def save_reports_to_excel(reports: List[EvaluationReport], output_path: Path) -> None:
    """Save evaluation reports to an Excel workbook with tidy + per-task overview sheets.

    Sheets:
        - "tidy": long-format, same content as CSV (via pandas).
        - One sheet per task: wide-format with multi-index columns
          (model_name merged + cfg keys as sub-rows).
    """
    import pandas as pd

    rows = _build_tidy_rows(reports)
    if not reports:
        return

    cfg_keys = sorted({k for r in reports if r.cfg_info for k in r.cfg_info.keys()})
    all_tasks = sorted({t.task_name for r in reports for t in r.task_results})
    fairness_keys = sorted(
        {key for r in reports for t in r.task_results for key in t.fairness.keys()}
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # Sheet tidy
        if rows:
            df_tidy = pd.DataFrame(rows)
            df_tidy.to_excel(writer, sheet_name="tidy", index=False)

        # One sheet per task
        for task_name in all_tasks:
            ws = writer.book.create_sheet(title=task_name)
            _write_overview_sheet(ws, reports, task_name, cfg_keys, fairness_keys)

    console.print(f"[green]Excel saved to {output_path}[/green]")
